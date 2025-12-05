from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.http import HttpResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.core.exceptions import ValidationError
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from .models import Vaquejada, Categoria, Ingresso
from django.conf import settings


# ==================== AUTENTICAÇÃO ====================

def home(request):
    vaquejadas = Vaquejada.objects.filter(status='open').order_by('-dia')[:6]
    context = {
        'vaquejadas': vaquejadas,
        'total_vaquejadas': Vaquejada.objects.filter(status='open').count(),
    }
    return render(request, 'core/home.html', context)


@require_http_methods(["GET", "POST"])
def register(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        
        if not username or not email or not password:
            messages.error(request, 'Todos os campos são obrigatórios.')
            return redirect('register')
        
        if password != password_confirm:
            messages.error(request, 'As senhas não coincidem.')
            return redirect('register')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Nome de usuário já existe.')
            return redirect('register')
        
        if User.objects.filter(email=email).exists():
            messages.error(request, 'E-mail já existe.')
            return redirect('register')
        
        if len(password) < 6:
            messages.error(request, 'A senha deve ter pelo menos 6 caracteres.')
            return redirect('register')
        
        user = User.objects.create_user(username=username, email=email, password=password)
        messages.success(request, 'Cadastro realizado com sucesso! Faça login para continuar.')
        return redirect('login')
    
    return render(request, 'core/register.html')


@require_http_methods(["GET", "POST"])
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, 'Login realizado com sucesso!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Usuário ou senha inválidos.')
    
    return render(request, 'core/login.html')


@login_required(login_url='login')
def logout_view(request):
    logout(request)
    messages.success(request, 'Logout realizado com sucesso!')
    return redirect('home')


# ==================== DASHBOARD ====================

@login_required(login_url='login')
def dashboard(request):
    total_ingressos = Ingresso.objects.filter(user=request.user).count()
    total_pagos = Ingresso.objects.filter(user=request.user, status='paid').count()
    total_pendentes = Ingresso.objects.filter(user=request.user, status='open').count()
    
    ingressos = Ingresso.objects.filter(user=request.user).order_by('-criado_em')
    
    context = {
        'total_ingressos': total_ingressos,
        'total_pagos': total_pagos,
        'total_pendentes': total_pendentes,
        'ingressos': ingressos,
    }
    return render(request, 'core/dashboard.html', context)


# ==================== VAQUEJADAS PÚBLICAS ====================

def vaquejadas_list(request):
    vaquejadas = Vaquejada.objects.filter(
        status='open'
    ).prefetch_related('categorias').order_by('-dia')
    
    local = request.GET.get('local', '').strip()
    if local:
        vaquejadas = vaquejadas.filter(local__icontains=local)
    
    context = {
        'vaquejadas': vaquejadas,
        'locais': Vaquejada.objects.filter(status='open').values_list('local', flat=True).distinct(),
    }
    return render(request, 'core/vaquejadas_list.html', context)


def vaquejada_detail(request, vaquejada_id):
    vaquejada = get_object_or_404(
        Vaquejada.objects.prefetch_related('categorias'),
        id=vaquejada_id
    )
    categorias = vaquejada.categorias.all()
    
    context = {
        'vaquejada': vaquejada,
        'categorias': categorias,
    }
    return render(request, 'core/vaquejada_detail.html', context)


@login_required(login_url='login')
def mapa_senhas(request, categoria_id):
    """Exibe o mapa de senhas disponíveis para uma categoria"""
    categoria = get_object_or_404(
        Categoria.objects.select_related('vaquejada'),
        id=categoria_id
    )
    
    # Buscar senhas ocupadas
    senhas_ocupadas = list(
        Ingresso.objects.filter(
            categoria=categoria,
            status__in=['open', 'closed', 'paid']
        ).values_list('numero_senha', flat=True)
    )
    
    # Criar mapa de senhas (1 até n_ingressos)
    mapa_senhas = []
    for numero in range(1, categoria.n_ingressos + 1):
        mapa_senhas.append({
            'numero': numero,
            'disponivel': numero not in senhas_ocupadas
        })
    
    context = {
        'categoria': categoria,
        'vaquejada': categoria.vaquejada,
        'mapa_senhas': mapa_senhas,
        'total_disponivel': len([s for s in mapa_senhas if s['disponivel']]),
        'total_vendidas': len(senhas_ocupadas),
    }
    return render(request, 'core/mapa_senhas.html', context)


# ==================== INGRESSOS ====================

@login_required(login_url='login')
@require_http_methods(["POST"])
def gerar_ingresso(request, categoria_id, numero_senha):
    """Gerar novo ingresso para uma categoria com número de senha específico"""
    categoria = get_object_or_404(Categoria, id=categoria_id)
    
    # Converter numero_senha para int
    try:
        numero_senha = int(numero_senha)
    except ValueError:
        messages.error(request, 'Número de senha inválido.')
        return redirect('mapa_senhas', categoria_id=categoria_id)
    
    # Validar número de senha
    if numero_senha < 1 or numero_senha > categoria.n_ingressos:
        messages.error(request, 'Número de senha inválido.')
        return redirect('mapa_senhas', categoria_id=categoria_id)
    
    # Verificar se senha já está ocupada
    if Ingresso.objects.filter(categoria=categoria, numero_senha=numero_senha).exists():
        messages.error(request, f'Senha {numero_senha} já está ocupada.')
        return redirect('mapa_senhas', categoria_id=categoria_id)
    
    # Criar ingresso
    ingresso = Ingresso.objects.create(
        categoria=categoria,
        numero_senha=numero_senha,
        user=request.user,
        representacao='',
        puxador='',
        esteiro='',
        cavalo_puxador='',
        cavalo_esteiro='',
    )
    
    messages.success(request, f'Senha {numero_senha} reservada com sucesso!')
    return redirect('ingresso_detail', ingresso_id=ingresso.id)


@login_required(login_url='login')
def meus_ingressos(request):
    ingressos = Ingresso.objects.filter(user=request.user).order_by('-criado_em')
    
    context = {
        'ingressos': ingressos,
    }
    return render(request, 'core/meus_ingressos.html', context)


@login_required(login_url='login')
def ingresso_detail(request, ingresso_id):
    ingresso = get_object_or_404(Ingresso, id=ingresso_id, user=request.user)
    
    context = {
        'ingresso': ingresso,
        'whatsapp_contact': settings.WHATSAPP_CONTACT,
    }
    return render(request, 'core/ingresso_detail.html', context)


@login_required(login_url='login')
@require_http_methods(["GET", "POST"])
def ingresso_edit(request, ingresso_id):
    ingresso = get_object_or_404(Ingresso, id=ingresso_id, user=request.user)
    
    if request.method == 'POST':
        representacao = request.POST.get('representacao', '').strip()
        puxador = request.POST.get('puxador', '').strip()
        esteiro = request.POST.get('esteiro', '').strip()
        cavalo_puxador = request.POST.get('cavalo_puxador', '').strip()
        cavalo_esteiro = request.POST.get('cavalo_esteiro', '').strip()
        
        if not all([representacao, puxador, esteiro, cavalo_puxador, cavalo_esteiro]):
            messages.error(request, 'Todos os campos são obrigatórios.')
            return redirect('ingresso_edit', ingresso_id=ingresso.id)
        
        ingresso.representacao = representacao
        ingresso.puxador = puxador
        ingresso.esteiro = esteiro
        ingresso.cavalo_puxador = cavalo_puxador
        ingresso.cavalo_esteiro = cavalo_esteiro
        ingresso.save()
        
        messages.success(request, 'Dados do ingresso atualizados com sucesso!')
        return redirect('ingresso_detail', ingresso_id=ingresso.id)
    
    context = {
        'ingresso': ingresso,
    }
    return render(request, 'core/ingresso_edit.html', context)


@login_required(login_url='login')
@require_http_methods(["POST"])
def enviar_comprovante_pix(request, ingresso_id):
    ingresso = get_object_or_404(Ingresso, id=ingresso_id, user=request.user)
    
    if 'comprovante_pix' not in request.FILES:
        messages.error(request, 'Nenhum arquivo foi enviado.')
        return redirect('ingresso_detail', ingresso_id=ingresso.id)
    
    arquivo = request.FILES['comprovante_pix']
    
    # Validar tamanho (máx 5MB)
    if arquivo.size > 5 * 1024 * 1024:
        messages.error(request, 'Arquivo muito grande. Máximo 5MB.')
        return redirect('ingresso_detail', ingresso_id=ingresso.id)
    
    # Validar tipo de arquivo
    tipos_permitidos = ['image/jpeg', 'image/png', 'image/jpg', 'application/pdf']
    if arquivo.content_type not in tipos_permitidos:
        messages.error(request, 'Tipo de arquivo não permitido. Use JPEG, PNG ou PDF.')
        return redirect('ingresso_detail', ingresso_id=ingresso.id)
    
    ingresso.comprovante_pix = arquivo
    ingresso.status = 'closed'  # Aguardando confirmação
    ingresso.pago_em = timezone.now()
    ingresso.save()
    
    messages.success(request, 'Comprovante enviado com sucesso! Aguarde a confirmação do organizador.')
    return redirect('ingresso_detail', ingresso_id=ingresso.id)


@login_required(login_url='login')
@require_http_methods(["POST"])
def admin_confirmar_pagamento(request, ingresso_id):
    """View para admin confirmar o pagamento de um ingresso"""
    if not request.user.is_staff:
        messages.error(request, 'Acesso negado.')
        return redirect('home')
    
    ingresso = get_object_or_404(Ingresso, id=ingresso_id)
    
    if ingresso.status == 'closed' and ingresso.comprovante_pix:
        ingresso.status = 'paid'
        ingresso.pago_em = timezone.now()
        ingresso.save()
        messages.success(request, f'Pagamento de {ingresso.representacao} confirmado!')
    else:
        messages.error(request, 'Ingresso não possui comprovante para confirmar.')
    
    return redirect('admin_ingressos', vaquejada_id=ingresso.categoria.vaquejada.id)


# ==================== ADMIN ====================

@login_required(login_url='login')
def admin_dashboard(request):
    if not request.user.is_staff:
        messages.error(request, 'Acesso negado.')
        return redirect('home')
    
    total_vaquejadas = Vaquejada.objects.count()
    total_ingressos = Ingresso.objects.count()
    total_vendido = Ingresso.objects.filter(status='paid').aggregate(total=Sum('categoria__valor'))['total'] or 0
    
    ingressos_pendentes = Ingresso.objects.filter(status='closed').count()
    
    context = {
        'total_vaquejadas': total_vaquejadas,
        'total_ingressos': total_ingressos,
        'total_vendido': total_vendido,
        'ingressos_pendentes': ingressos_pendentes,
    }
    return render(request, 'core/admin_dashboard.html', context)


@login_required(login_url='login')
def admin_vaquejadas(request):
    if not request.user.is_staff:
        messages.error(request, 'Acesso negado.')
        return redirect('home')
    
    vaquejadas = Vaquejada.objects.all().order_by('-criado_em')
    
    context = {
        'vaquejadas': vaquejadas,
    }
    return render(request, 'core/admin_vaquejadas.html', context)


@login_required(login_url='login')
@require_http_methods(["GET", "POST"])
def admin_vaquejada_edit(request, vaquejada_id=None):
    if not request.user.is_staff:
        messages.error(request, 'Acesso negado.')
        return redirect('home')
    
    vaquejada = None
    if vaquejada_id:
        vaquejada = get_object_or_404(Vaquejada, id=vaquejada_id)
    
    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        descricao = request.POST.get('descricao', '').strip()
        dia = request.POST.get('dia', '').strip()
        hora = request.POST.get('hora', '').strip()
        local = request.POST.get('local', '').strip()
        status = request.POST.get('status', 'open').strip()
        
        if not all([nome, descricao, dia, hora, local]):
            messages.error(request, 'Todos os campos obrigatórios devem ser preenchidos.')
            return redirect('admin_vaquejada_edit', vaquejada_id=vaquejada_id)
        
        if vaquejada:
            vaquejada.nome = nome
            vaquejada.descricao = descricao
            vaquejada.dia = dia
            vaquejada.hora = hora
            vaquejada.local = local
            vaquejada.status = status
            if 'imagem' in request.FILES:
                vaquejada.imagem = request.FILES['imagem']
            vaquejada.save()
            messages.success(request, 'Vaquejada atualizada com sucesso!')
        else:
            vaquejada = Vaquejada.objects.create(
                nome=nome,
                descricao=descricao,
                dia=dia,
                hora=hora,
                local=local,
                status=status,
                imagem=request.FILES.get('imagem', None),
            )
            messages.success(request, 'Vaquejada criada com sucesso!')
        
        return redirect('admin_vaquejadas')
    
    context = {
        'vaquejada': vaquejada,
    }
    return render(request, 'core/admin_vaquejada_edit.html', context)


@login_required(login_url='login')
def admin_ingressos(request, vaquejada_id):
    if not request.user.is_staff:
        messages.error(request, 'Acesso negado.')
        return redirect('home')
    
    vaquejada = get_object_or_404(Vaquejada, id=vaquejada_id)
    ingressos = Ingresso.objects.filter(categoria__vaquejada=vaquejada).order_by('-criado_em')
    
    context = {
        'vaquejada': vaquejada,
        'ingressos': ingressos,
    }
    return render(request, 'core/admin_ingressos.html', context)


@login_required(login_url='login')
def admin_relatorio(request, vaquejada_id):
    if not request.user.is_staff:
        messages.error(request, 'Acesso negado.')
        return redirect('home')
    
    vaquejada = get_object_or_404(Vaquejada, id=vaquejada_id)
    categorias = vaquejada.categorias.all()
    
    data = []
    total_geral = 0
    for categoria in categorias:
        ingressos = Ingresso.objects.filter(categoria=categoria)
        pago = ingressos.filter(status='paid').count()
        pendente = ingressos.filter(status__in=['open', 'closed']).count()
        valor_total = float(categoria.valor) * ingressos.count()
        
        data.append({
            'categoria': categoria.nome,
            'quantidade': ingressos.count(),
            'pago': pago,
            'pendente': pendente,
            'valor_total': valor_total,
        })
        total_geral += valor_total
    
    context = {
        'vaquejada': vaquejada,
        'data': data,
        'total_geral': total_geral,
    }
    return render(request, 'core/admin_relatorio.html', context)


@login_required(login_url='login')
def exportar_excel(request, vaquejada_id):
    if not request.user.is_staff:
        messages.error(request, 'Acesso negado.')
        return redirect('home')
    
    vaquejada = get_object_or_404(Vaquejada, id=vaquejada_id)
    ingressos = Ingresso.objects.filter(categoria__vaquejada=vaquejada)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ingressos'
    
    headers = ['Representação', 'Puxador', 'Esteiro', 'Cavalo Puxador', 'Cavalo Esteiro', 'Categoria', 'Status', 'Data Criação']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    for row, ingresso in enumerate(ingressos, 2):
        ws.cell(row=row, column=1).value = ingresso.representacao
        ws.cell(row=row, column=2).value = ingresso.puxador
        ws.cell(row=row, column=3).value = ingresso.esteiro
        ws.cell(row=row, column=4).value = ingresso.cavalo_puxador
        ws.cell(row=row, column=5).value = ingresso.cavalo_esteiro
        ws.cell(row=row, column=6).value = ingresso.categoria.nome
        ws.cell(row=row, column=7).value = ingresso.get_status_display()
        ws.cell(row=row, column=8).value = ingresso.criado_em.strftime('%d/%m/%Y %H:%M')
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ingressos_{vaquejada.nome.replace(" ", "_")}.xlsx"'
    wb.save(response)
    
    return response


@login_required(login_url='login')
def exportar_csv(request, vaquejada_id):
    if not request.user.is_staff:
        messages.error(request, 'Acesso negado.')
        return redirect('home')
    
    vaquejada = get_object_or_404(Vaquejada, id=vaquejada_id)
    ingressos = Ingresso.objects.filter(categoria__vaquejada=vaquejada)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="ingressos_{vaquejada.nome.replace(" ", "_")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Representação', 'Puxador', 'Esteiro', 'Cavalo Puxador', 'Cavalo Esteiro', 'Categoria', 'Status', 'Data'])
    
    for ingresso in ingressos:
        writer.writerow([
            ingresso.representacao,
            ingresso.puxador,
            ingresso.esteiro,
            ingresso.cavalo_puxador,
            ingresso.cavalo_esteiro,
            ingresso.categoria.nome,
            ingresso.get_status_display(),
            ingresso.criado_em.strftime('%d/%m/%Y %H:%M'),
        ])
    
    return response
